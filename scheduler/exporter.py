"""
Import / Export 모듈.

38종 근무 코드 대응:
  - SHIFT_META 로 셀 색상 자동 결정
  - 코드 값 그대로 Excel/CSV 에 기재 (D, E, N, N7, Y, I, T 등)
  - ERP 레코드에 shift_label(한국어 명칭) 추가
"""

from __future__ import annotations

import csv
import datetime
import io
import json
from typing import Dict, List, Optional

import pandas as pd

from .models import (
    SHIFT_META,
    WORK_SHIFTS,
    Nurse,
    Schedule,
    ScheduleConfig,
    ScheduleEntry,
    ShiftType,
    get_shift_label,
)

WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

# OFF / 미배정 기본색
_DEFAULT_COLOR = "FFFFFF"


def _cell_color(shift: ShiftType) -> str:
    meta = SHIFT_META.get(shift)
    return meta.color_hex if meta else _DEFAULT_COLOR


class ScheduleExporter:
    """근무표 Import / Export."""

    def __init__(self, nurses: List[Nurse]) -> None:
        self.nurses = nurses
        self.nurse_map = {n.id: n for n in nurses}

    # ── DataFrame ─────────────────────────────

    def to_dataframe(self, schedule: Schedule) -> pd.DataFrame:
        """Schedule → DataFrame (행=간호사, 열=날짜, 값=코드)."""
        matrix = schedule.as_matrix(self.nurses)
        dates = sorted({e.date for e in schedule.entries})

        rows = []
        for nurse in self.nurses:
            row = {"간호사": f"{nurse.name}({nurse.skill_level.value})", "ID": nurse.id}
            for d in dates:
                col = f"{d.month}/{d.day}\n{WEEKDAY_KR[d.weekday()]}"
                row[col] = matrix[nurse.id].get(d, ShiftType.O).value
            rows.append(row)

        df = pd.DataFrame(rows).set_index("간호사")
        df.drop(columns=["ID"], inplace=True)
        return df

    def to_summary_dataframe(self, schedule: Schedule) -> pd.DataFrame:
        """간호사별 근무 통계 요약 DataFrame."""
        from .models import NIGHT_SHIFTS, OFF_SHIFTS
        matrix = schedule.as_matrix(self.nurses)
        rows = []
        for nurse in self.nurses:
            hist = matrix[nurse.id]
            work = sum(1 for s in hist.values() if s in WORK_SHIFTS)
            nights = sum(1 for s in hist.values() if s in NIGHT_SHIFTS)
            weekends = sum(1 for d, s in hist.items() if d.weekday() >= 5 and s in WORK_SHIFTS)
            # 주요 코드 카운트
            counts = {}
            for s in hist.values():
                counts[s.value] = counts.get(s.value, 0) + 1
            rows.append({
                "이름": nurse.name,
                "경력": nurse.skill_level.value,
                "근무일": work,
                "D": counts.get("D", 0),
                "E": counts.get("E", 0),
                "N": counts.get("N", 0),
                "N7": counts.get("N7", 0),
                "M": counts.get("M", 0),
                "야간합계": nights,
                "주말": weekends,
                "연차(Y)": counts.get("Y", 0),
                "병가(I)": counts.get("I", 0),
                "교육(T)": counts.get("T", 0),
                "비번(O)": counts.get("O", 0) + counts.get("OFF", 0),
            })
        return pd.DataFrame(rows)

    # ── CSV ───────────────────────────────────

    def to_csv(self, schedule: Schedule) -> str:
        return self.to_dataframe(schedule).to_csv(encoding="utf-8-sig")

    def from_csv(self, csv_content: str, ward_id: str, year: int, month: int) -> Schedule:
        df = pd.read_csv(io.StringIO(csv_content), index_col=0)
        return self._dataframe_to_schedule(df, ward_id, year, month)

    # ── Excel ─────────────────────────────────

    def to_excel(self, schedule: Schedule) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = f"{schedule.year}-{schedule.month:02d} 근무표"

        matrix = schedule.as_matrix(self.nurses)
        dates = sorted({e.date for e in schedule.entries})
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 헤더 행
        ws.cell(row=1, column=1, value="간호사").font = Font(bold=True)
        for ci, d in enumerate(dates, start=2):
            label = f"{d.month}/{d.day}({WEEKDAY_KR[d.weekday()]})"
            cell = ws.cell(row=1, column=ci, value=label)
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal="center")
            if d.weekday() == 5:                          # 토요일
                cell.fill = PatternFill("solid", fgColor="CCE5FF")
            elif d.weekday() == 6:                        # 일요일
                cell.fill = PatternFill("solid", fgColor="FFE5E5")
            cell.border = border

        # 데이터 행
        for ri, nurse in enumerate(self.nurses, start=2):
            name_cell = ws.cell(
                row=ri, column=1,
                value=f"{nurse.name} ({nurse.skill_level.value})"
            )
            name_cell.font = Font(bold=(nurse.skill_level.value == "숙련"), size=9)
            name_cell.border = border

            for ci, d in enumerate(dates, start=2):
                shift = matrix[nurse.id].get(d, ShiftType.O)
                cell = ws.cell(row=ri, column=ci, value=shift.value)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill("solid", fgColor=_cell_color(shift))
                cell.font = Font(size=9, bold=(shift in WORK_SHIFTS))
                cell.border = border

        # 열 너비
        ws.column_dimensions["A"].width = 16
        for ci in range(2, len(dates) + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 5.5
        ws.row_dimensions[1].height = 28

        # 근무 코드 범례 시트
        ws2 = wb.create_sheet("근무코드표")
        ws2.append(["코드", "명칭", "카테고리", "근무여부"])
        from .models import SHIFT_META, ShiftType as ST
        for s, meta in SHIFT_META.items():
            r = ws2.max_row + 1
            ws2.cell(row=r, column=1, value=s.value)
            ws2.cell(row=r, column=2, value=meta.label)
            ws2.cell(row=r, column=3, value=meta.category)
            ws2.cell(row=r, column=4, value="○" if meta.is_work else "—")
            ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=meta.color_hex)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def from_excel(self, excel_bytes: bytes, ward_id: str, year: int, month: int) -> Schedule:
        df = pd.read_excel(io.BytesIO(excel_bytes), index_col=0, sheet_name=0)
        return self._dataframe_to_schedule(df, ward_id, year, month)

    # ── JSON ──────────────────────────────────

    def to_image(self, schedule: Schedule) -> bytes:
        """
        Schedule → 색상 표현 PNG 이미지 (matplotlib).
        각 셀을 SHIFT_META 색상으로 채운 표를 렌더링한다.
        """
        import io as _io

        import matplotlib
        matplotlib.use("Agg")  # 헤드리스 환경에서 GUI 백엔드 불필요
        import matplotlib.pyplot as plt

        df = self.to_dataframe(schedule)
        n_nurses = len(df)
        n_days = len(df.columns)

        fig_w = max(14, n_days * 0.48 + 3.0)
        fig_h = max(4, n_nurses * 0.42 + 1.8)

        fig, ax = plt.subplots(figsize=(fig_w, fig_h))
        ax.set_axis_off()
        fig.patch.set_facecolor("#F1F4FA")

        col_labels = list(df.columns)
        row_labels = [str(r) for r in df.index]
        cell_text = [
            [str(v) for v in row]
            for row in df.values.tolist()
        ]

        # 셀 배경색: ShiftType → (r, g, b, alpha)
        def _hex_to_rgba(hex6: str, alpha: float = 0.80):
            h = hex6.lstrip("#")
            return (int(h[0:2], 16) / 255, int(h[2:4], 16) / 255, int(h[4:6], 16) / 255, alpha)

        cell_colors = []
        for row_data in cell_text:
            row_colors = []
            for val in row_data:
                try:
                    shift = ShiftType(val)
                    hex_col = "#" + _cell_color(shift)
                except ValueError:
                    hex_col = "#FFFFFF"
                row_colors.append(_hex_to_rgba(hex_col))
            cell_colors.append(row_colors)

        tbl = ax.table(
            cellText=cell_text,
            rowLabels=row_labels,
            colLabels=col_labels,
            cellColours=cell_colors,
            loc="center",
            cellLoc="center",
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(7.5)
        tbl.scale(1.0, 1.35)

        # 헤더 행 스타일 (navy)
        for j in range(n_days):
            cell = tbl[(0, j)]
            cell.set_facecolor("#1B2238")
            cell.set_text_props(color="white", fontweight="bold", fontsize=7)

        # 행 레이블 스타일
        for i in range(n_nurses):
            cell = tbl[(i + 1, -1)]
            cell.set_facecolor("#26314F")
            cell.set_text_props(color="white", fontweight="bold", fontsize=8)

        ax.set_title(
            f"{schedule.year}년 {schedule.month}월 근무표",
            pad=12, fontsize=13, fontweight="bold", color="#1B2238",
        )
        plt.tight_layout(pad=0.3)

        buf = _io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        return buf.getvalue()

    def to_json(self, schedule: Schedule) -> str:
        return schedule.model_dump_json(indent=2)

    def from_json(self, json_str: str) -> Schedule:
        return Schedule.model_validate_json(json_str)

    # ── ERP / OCS 연동 ────────────────────────

    def export_for_erp(self, schedule: Schedule) -> List[Dict]:
        """표준 ERP 레코드 목록 반환."""
        records = []
        for entry in schedule.entries:
            nurse = self.nurse_map.get(entry.nurse_id)
            records.append({
                "nurse_id":    entry.nurse_id,
                "nurse_name":  nurse.name if nurse else "",
                "date":        entry.date.isoformat(),
                "shift_code":  entry.shift.value,
                "shift_label": get_shift_label(entry.shift),
                "is_work":     entry.shift in WORK_SHIFTS,
                "is_holiday":  entry.is_holiday,
                "is_weekend":  entry.is_weekend,
                "is_fixed":    entry.is_fixed,
            })
        return records

    def import_from_erp(self, records: List[Dict], ward_id: str, year: int, month: int) -> Schedule:
        entries = []
        for rec in records:
            try:
                shift = ShiftType(rec["shift_code"])
            except ValueError:
                shift = ShiftType.O
            entries.append(ScheduleEntry(
                nurse_id=rec["nurse_id"],
                date=datetime.date.fromisoformat(rec["date"]),
                shift=shift,
                is_fixed=rec.get("is_fixed", False),
                is_holiday=rec.get("is_holiday", False),
                is_weekend=rec.get("is_weekend", False),
            ))
        return Schedule(ward_id=ward_id, year=year, month=month, entries=entries,
                        generated_at=datetime.datetime.now())

    # ── 내부 헬퍼 ─────────────────────────────

    def _dataframe_to_schedule(
        self, df: pd.DataFrame, ward_id: str, year: int, month: int
    ) -> Schedule:
        entries = []
        name_to_id = {f"{n.name}({n.skill_level.value})": n.id for n in self.nurses}

        for nurse_label in df.index:
            nurse_id = name_to_id.get(str(nurse_label))
            if nurse_id is None:
                # 이름만으로 fallback 매칭
                for n in self.nurses:
                    if n.name in str(nurse_label):
                        nurse_id = n.id
                        break
            if nurse_id is None:
                continue

            for col in df.columns:
                try:
                    date_part = str(col).split("\n")[0].split("(")[0].strip()
                    m, d = map(int, date_part.split("/"))
                    date = datetime.date(year, m, d)
                except (ValueError, IndexError):
                    continue
                raw = str(df.loc[nurse_label, col]).strip()
                try:
                    shift = ShiftType(raw)
                except ValueError:
                    shift = ShiftType.O
                entries.append(ScheduleEntry(
                    nurse_id=nurse_id, date=date, shift=shift,
                    is_fixed=True, is_weekend=date.weekday() >= 5,
                ))
        return Schedule(ward_id=ward_id, year=year, month=month, entries=entries,
                        generated_at=datetime.datetime.now())
